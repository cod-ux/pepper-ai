import os

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

import xlwings as xl

from PoC_v3 import (
    create_plan,
    generate_code,
    review_code,
    refresh_code
)


if 'key' not in st.session_state:
    st.session_state.key = 'value'

if "messages" not in st.session_state.keys():
    st.session_state.messages = [{
        "role": "ai", 
        "content": ""
        }]

def copy_excel_locally(file):
    fname, ext = os.path.splitext(file.name)
    new_fname = f"{fname}_copy{ext}"
    file_root = os.path.join("/Users/suryaganesan/vscode/ml/projects/reporter/uploads", new_fname)
    with open(file_root, "wb") as local_file:
        local_file.write(file.read())
    #st.success(f"File saved as {new_fname}")

    file_path = f"/Users/suryaganesan/vscode/ml/projects/reporter/uploads/{new_fname}"

    return file_path

@st.cache_data()
def load_sheets_to_dfs(file_path):
    wb = load_workbook(filename=file_path, data_only=False)
    dfs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = ws.values
        columns = next(data)
        data = list(data)
        dfs.append(pd.DataFrame(data, columns=columns))
    return dfs, wb.sheetnames

def reload_sheets_to_dfs(file_path):
    wb = load_workbook(filename=file_path, data_only=False)
    dfs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = ws.values
        columns = next(data)
        data = list(data)
        dfs.append(pd.DataFrame(data, columns=columns))
    return dfs, wb.sheetnames

def save_sheets(path):
    app = xl.App(visible=False)
    book = xl.Book(path)
    book.save()
    book.close()
    app.kill()
    app.quit()
    return True

# Main function
def main():
    # Upload Excel file
    header_ph = st.empty()
    header_ph.markdown( "<h1 style='text-align: center;'>Command AI: Clean & prepare your excel faster</h1>", unsafe_allow_html=True)
    uploader_ph = st.empty()

    with uploader_ph.container():
        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])

    if uploaded_file is not None:

        file_path = copy_excel_locally(uploaded_file)
        
        if "file_path" not in st.session_state:
           file_path = copy_excel_locally(uploaded_file)
           st.session_state.file_path = file_path

        else:
            file_path = st.session_state.file_path 

        # Save excel file
        st.text("Current State: ")
        dfs, sheets = load_sheets_to_dfs(file_path)

        select_sheet_ph = st.empty()

        with select_sheet_ph.container():
            current_sheet = st.selectbox(
                "Select sheet",
                (sheets)
            )

        # Display current state
        current_table_placeholder = st.empty()

        with current_table_placeholder.container():
            st.dataframe(dfs[sheets.index(current_sheet)])

        request = st.chat_input("Enter your command...")
        
        if request:
            st.session_state.messages.append({"role": "human", "content": request})
            with st.chat_message("user"):
                st.write(request)

        if st.session_state.messages[-1]["role"] != "ai":
            if request:
               with st.chat_message("ai"):
                 if request != "quit":
                    st.write("Creating a plan...")
                    plan = create_plan(request, file_path)

                    st.write(f"The plan: \n{plan}")

                    st.write("Generating script")
                    script_generated = generate_code(request, file_path, plan)

                    st.code(f"Code generated: \n {script_generated}")
                    print("\n\nReviewing code...")

                    reviewed_script = review_code(request, script_generated, file_path, plan)

                    final_script = reviewed_script

                    st.code(f"#Script reviewed: \n{final_script}")

                    try:
                        st.write("\n\nInitiating code execution...")
                        exec(final_script)

                    except Exception as e:
                        st.write(f"Error: {e}")
                        st.write("Analysing the error....")
                        new_code = refresh_code(request, e, file_path, plan, final_script)
                        st.code(f"New code: \n {new_code}")

                        try:
                            st.write("Initializing new refreshed code....")
                            exec(new_code)

                        except Exception as e2:
                            st.write(f"New Error: {e2}")

                
                    trigger = save_sheets(file_path)
                    if trigger:
                        st.cache_data.clear()

                        dfs, sheets = load_sheets_to_dfs(file_path)

                        current_table_placeholder.empty()
                        select_sheet_ph.empty()
                        uploader_ph.empty()

                        with select_sheet_ph.container():
                            current_sheet = st.selectbox(
                                "Select sheets",
                                (sheets)
                            )

                        with current_table_placeholder.container():
                            st.dataframe(dfs[sheets.index(current_sheet)])

                        trigger = False




        # Display current state
        

        

        





    # Create buttons for each sheet
    # if st.button show df for each sheet



if __name__ == "__main__":
    main()
