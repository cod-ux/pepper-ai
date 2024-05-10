from openpyxl import load_workbook
import json
import os
import toml

from openai import OpenAI


# Setup

source = "sales_data_sample.xlsx"
secrets = "/Users/suryaganesan/Documents/GitHub/Replicate/secrets.toml"
os.environ["OPENAI_API_KEY"] = toml.load(secrets)["OPENAI_API_KEY"]

client = OpenAI()



def query_llm(user_msg, client=client):
    system_msg = f"""You are an AI chatbot that will be given the name of an excel file, a JSON representation of the excel file and some instructions from the user on how to manipulate the excel file.

    Write python code with the openpyxl library that can be used to execute the user's commands on the excel file.

    Load the excel file, execute the user commands, convert it back to excel and save it. Prefer to use formulas over manual data entry.
    
    Provide error free PYTHON CODE that I will dynamically execute as part of another program with the exec() function. Provide ONLY the CODE.
    """

    messages = [
        {
            "role": "system",
            "content": system_msg
        }
    ]

    messages.append(
        {
            "role": "user",
            "content": user_msg
        }
    )

    response = client.chat.completions.create(
        model="gpt-4",
        messages=messages
    ).choices[0].message

    return response

def load_excel_to_json(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=False)
    sheet = workbook.active

    # Convert Excel data to a list of dictionaries
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
           data.append(row)

    # Convert list of dictionaries to JSON
    jstring = json.dumps(data)

    return jstring


def save_last_as_json(json_text):
    with open("last_save.json", "w") as json_file:
        data = json.loads(json_text)
        json.dump(data, json_file, indent=4)

def extract_code(content):
    start_index = content.find("""```python""") + len("""'''python""")
    end_index = content.find("""```""", start_index)

    python_code = content[start_index:end_index].strip()
    return python_code

print("Program begins....\n\n")

