from openpyxl import load_workbook
import json
import os
import toml

from openai import OpenAI
import pandas as pd


# Setup

source = "sales_data_sample.xlsx"
secrets = "/Github/reporter/secrets.toml"
os.environ["OPENAI_API_KEY"] = toml.load(secrets)["OPENAI_API_KEY"]

client = OpenAI()

pd.set_option('display.max_columns', None)
pd.set_option('display.expand_frame_repr', False)

def query_llm_gpt4(user_msg, client=client):
    system_msg = f"""You are an AI python programmer. You are adept in writing and reviewing code. Assume that openpyxl library is already installed when writing code.
    """

    messages = [
        {
            "role": "system",
            "content": system_msg
        }
    ]

    messages.append(
        {
            "role": "user",
            "content": user_msg
        }
    )

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=messages
    ).choices[0].message

    return response

def query_llm_gpt35(user_msg, client=client):
    system_msg = f"""You are an AI python programmer. You are adept in writing and reviewing code. Assume that openpyxl library is already installed when writing code.
    """

    messages = [
        {
            "role": "system",
            "content": system_msg
        }
    ]

    messages.append(
        {
            "role": "user",
            "content": user_msg
        }
    )

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=messages
    ).choices[0].message

    return response

def load_excel_to_json(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=False)
    sheet = workbook.active

    # Convert Excel data to a list of dictionaries
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
           data.append(row)

    # Convert list of dictionaries to JSON
    jstring = json.dumps(data)

    return jstring

def load_excel_to_df(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=False)
    sheet = workbook.active

    # Convert Excel data to a list of dictionaries
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
           data.append(row)

    df = pd.DataFrame(data[1:], columns=data[0])

    return df

def load_sheets_to_dfs(file_path):
    wb = load_workbook(filename=file_path, data_only=False)
    dfs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = ws.values
        try:
            columns = next(data)
            data = list(data)
            if data:
                dfs.append(pd.DataFrame(data, columns=columns))
            else:
                dfs.append(pd.DataFrame())
        
        except StopIteration:
            dfs.append(pd.DataFrame())
    return dfs, wb.sheetnames

def save_last_as_json(json_text):
    with open("last_save.json", "w") as json_file:
        data = json.loads(json_text)
        json.dump(data, json_file, indent=4)

def extract_code_from_llm(content):
    start_index = content.find("""```python""") + len("""```python""")
    end_index = content.find("""```""", start_index)

    python_code = content[start_index:end_index].strip()
    return python_code


