import xlwings as xl
import pandas as pd
from openpyxl import load_workbook
import json
import os
import toml

from openai import OpenAI

print("Program begins....\n\n")

source = "sales_data_sample.xlsx"
secrets = "C:/Users/Administrator/Documents/reporter/secrets.toml"
os.environ["OPENAI_API_KEY"] = toml.load(secrets)["OPENAI_API_KEY"]

client = OpenAI()

def query_llm(user_msg, client=client):
    system_msg = """You are an AI chatbot that takes in json representations of excel files and user instructions on how to modify the excel file.
    You will execute user requests through editing the json representation of the excel file by editing, modifying or creating new columns, rows or cells.
    Always prefer to write formulas for data manipulation related requests like creating a new column based on conditions relating to pre-existing columns.
    Use null to represent empty values and not None to comply with JSON format rules, an ONLY USE DOUBLE QUOETS. DO NOT USE SINGLE QUOTES.
    Return ONLY a JSON object as response WITHOUT any describing words. Respond in JSON format."""
    messages = [
        {
            "role": "system",
            "content": system_msg
        }
    ]

    messages.append(
        {
            "role": "user",
            "content": user_msg
        }
    )

    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=messages
    ).choices[0].message

    return response

def excel_to_json(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=False)
    sheet = workbook.active

    # Convert Excel data to a list of dictionaries
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
           data.append(row)

    # Convert list of dictionaries to JSON
    jstring = json.dumps(data)

    return jstring

def save_last_json(json_text):
    with open("last_save.json", "w") as json_file:
        data = json.loads(json_text)
        json.dump(data, json_file, indent=4)

def json_to_excel(json_data):
    wb = load_workbook("save.xlsx")
    sheet = wb.active

    for row_id, row_data in enumerate(json_data, start=1):
        for col_idx, cell_data in enumerate(row_data, start=1):
            sheet.cell(row=row_id, column=col_idx, value=cell_data)

    wb.save("save.xlsx")

get_path = True
jtext = None

while get_path:
    try:
        wait = input(f"Excel source: {source}\nProceeding on confirmation...")
        jtext = excel_to_json(source)
        json_obj = json.loads(jtext)[:20]
        for i in json_obj:
            print(i, '\n\n')
        jtext = json.dumps(json_obj)
        save_last_json(jtext)
        print("\nLength of jtext: ", len(jtext))
        print(jtext)
        get_path = False

    except Exception as e:
        print("Error: ", e)


kill_cli = False
print("--------------- COMMAND LINE -----------------")

while not kill_cli:
    request = input(">> ")
    with open("last_save.json", "r") as last_save:
        jtext = json.load(last_save)
    if request:
        if request != "quit" and request != "save":
            prompt = f"""JSON representation of Excel file: \n{jtext}\n\nInstruction: {request}"""
            json_response = query_llm(prompt)
            print(json_response.content)
            save_last_json(json_response.content)

        elif request == "save":
            with open("last_save.json", "w") as last_save:
                data = json.load(last_save)
                json_to_excel(data)

        else:
            kill_cli = True



print("---------------EOP---------------")

# Path = /users/suryaganesan/downloads/test_book.xlsx

# Program starts
# Asks excel path
# Converts excel to json
# Ask for change
# Sends json to LLM to get back changes
# Print json