import streamlit as st
import pandas as pd
from openpyxl import load_workbook

import xlwings as xl

def unmerge_sheets(file_path):
    book = load_workbook(file_path)
    for sheet in book.sheetnames:
        ws = book[sheet]
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
            print("Merged Range: ", merged_range)

    book.save(file_path)
    book.close()
