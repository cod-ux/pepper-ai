from openpyxl import load_workbook
import json
import os
import toml

from openai import OpenAI
import pandas as pd


# Setup

source = "sales_data_sample.xlsx"
secrets = "C:/Users/Administrator/Documents/github/reporter/secrets.toml"
os.environ["OPENAI_API_KEY"] = toml.load(secrets)["OPENAI_API_KEY"]

client = OpenAI()

pd.set_option('display.max_columns', None)
pd.set_option('display.expand_frame_repr', False)

def query_llm_gpt4_plan(user_msg, client=client):
    messages = []

    messages.append(
        {
            "role": "system",
            "content": "You are an expert in understanding and clarifying requests from the user regarding excel operations. If the user request is related to creating new cells and columns, prefer to write excel formulas through the python code"
        }
    )

    messages.append(
        {
            "role": "user",
            "content": user_msg
        }
    )

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=messages
    ).choices[0].message

    return response

def query_llm_gpt4_code(user_msg, client=client):
    messages = []

    messages.append(
        {
            "role": "system",
            "content": """
You are an expert in writing python code from natural language requests to execute excel tasks. \n
You write python code to execute the user rquest eith the openpyxl library. \n
You can also try to write excel formulas in the sheets with your code when the request is related to creating new columns and it is simpler to do so than writing code.\n
ONLY use openpyxl module for loading excel data.
Provide ONLY the python source code as a response.
Import ALL Necessary libraries in the code.
"""
        }
    )

    messages.append(
        {
            "role": "user",
            "content": user_msg
        }
    )

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=messages
    ).choices[0].message

    return response

def query_llm_gpt35(user_msg, client=client):
    prompt = f"""
{user_msg}
"""

    response = client.completions.create(
        model="gpt-3.5-turbo-instruct",
        prompt=prompt,
        max_tokens=1096,
        temperature=0.1
    ).choices[0].text

    return response

def query_llm_gpt35_chat(user_msg, client=client):
    prompt = f"""
{user_msg}
"""
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[{"role": "user", "content": prompt}],
        temperature=0
    ).choices[0].message.content

    return response

def load_excel_to_json(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=False)
    sheet = workbook.active

    # Convert Excel data to a list of dictionaries
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
           data.append(row)

    # Convert list of dictionaries to JSON
    jstring = json.dumps(data)

    return jstring

def load_excel_to_df(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=False)
    sheet = workbook.active

    # Convert Excel data to a list of dictionaries
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
           data.append(row)

    df = pd.DataFrame(data[1:], columns=data[0])

    return df

def load_sheets_to_dfs(file_path):
    wb = load_workbook(filename=file_path, data_only=False)
    dfs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = ws.values
        try:
            columns = next(data)
            data = list(data)
            if data:
                dfs.append(pd.DataFrame(data, columns=columns))
            else:
                dfs.append(pd.DataFrame())
        
        except StopIteration:
            dfs.append(pd.DataFrame())
    return dfs, wb.sheetnames

def save_last_as_json(json_text):
    with open("last_save.json", "w") as json_file:
        data = json.loads(json_text)
        json.dump(data, json_file, indent=4)

def extract_code_from_llm(content):
    start_index = content.find("""```python""") + len("""```python""")
    end_index = content.find("""```""", start_index)

    python_code = content[start_index:end_index].strip()
    return python_code


"""
You are a python programming assistant that writes code to execute user requests on their excel data using the openpyxl module. \n
Your goal is to write code that when executed achieves the user request on the given excel file. \n
Use the sample data provided from the excel sheets to identify necessary variables or write your code. \n

{user_msg}

Provide ONLY the code that is readily executable and not any other words or responses. \n


Write the code to execute the task:
### CODE
"""