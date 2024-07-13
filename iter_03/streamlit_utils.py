import os 
import xlwings as xl
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import time, random, re

def re_upload():
    st.session_state.uploaded_file = None
    st.session_state.file_path = None
    st.cache_data.clear()

def generate_unique_filename(directory, fname, ext):
    while True:
        timestamp = int(time.time())
        random_num = random.randint(1000, 9999)
        new_fname = f"{fname}_{timestamp}_{random_num}{ext}"
        file_path = f"{directory}/{new_fname}"
        if not os.path.exists(file_path):
            return new_fname

def copy_excel_locally(file):

    directory = "C:/Users/Administrator/Documents/reporter/uploads"
    os.makedirs(directory, exist_ok=True)
    
    fname, ext = os.path.splitext(file.name)
    new_fname = generate_unique_filename(directory, fname, ext)
    
    file_root = f"{directory}/{new_fname}"
    
    # Log the constructed file path for debugging
    print(f"Saving file to: {file_root}")
    
    with open(file_root, "wb") as local_file:
        local_file.write(file.read())

    file_path = file_root

    return file_path

def copy_excel_locally_from_path(file):
    fname, ext = os.path.splitext(file)
    new_fname = f"{fname}_copy{ext}"
    file_root = os.path.join("C:/Users/Administrator/Documents/reporter/uploads", new_fname)
    with open(file_root, "wb") as local_file:
        local_file.write(file.read())
    st.success(f"File saved as {new_fname}")

    file_path = f"C:/Users/Administrator/Documents/reporter/uploads/{new_fname}"

    return file_path


def handle_duplicate_columns(columns):
    counts = {}
    new_columns = []

    for col_name in columns:
        if col_name in counts:
            counts[col_name] += 1
            new_columns.append(f"{col_name}_{counts[col_name]}")
        else:
            counts[col_name] = 0
            new_columns.append(col_name)
    
    return new_columns


@st.cache_data()
def load_sheets_to_dfs(file_path):
    app = xl.App(visible=False)
    wb = app.books.open(file_path)
    dfs = []
    for sheet in wb.sheets:
        df = sheet.used_range.options(pd.DataFrame, header=True, index=False).value
        df.columns = handle_duplicate_columns(df.columns)
        dfs.append(df)

    sheet_names = [sheet.name for sheet in wb.sheets]
    wb.save()
    wb.close()
    app.quit()
    return dfs, sheet_names


def save_sheets(path):
    app = xl.App(visible=False)
    book = app.books.open(path)
    book.save()
    book.close()
    app.kill()
    return True

def unmerge_sheets(file_path):
    book = load_workbook(file_path)
    for sheet in book.sheetnames:
        ws = book[sheet]
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged_range))
                print("Merged Range: ", merged_range)

    book.save(file_path)
    book.close()

def get_binary(file_path):
    with open(file_path, "rb") as f:
        bn = f.read()
        
    return bn

def undo(file_path):
    try:
        last_state = st.session_state.state_stack.pop()
        if last_state:
            with open(file_path, "wb") as f:
                f.write(last_state)

            st.cache_data.clear()

    except Exception as e:
        print("Error with undo action: ", e)
        return

def split_numbered_list(text):
    # Regex to identify numbered list patterns specifically in the form "1. ", "2. ", etc.
    pattern = re.compile(r'(?<!\d)(?:\d+\.\s)(.*?)(?=\d+\.\s|$)', re.DOTALL)
    
    # Find all matches using the pattern
    matches = pattern.findall(text)
    
    # If matches are found, return the list of matches; else return the original text as a single-item list
    return matches if matches else [text]

"""
1. Fill empty cells with N/A, delete the new sheet
2. Copy all deals whose Deal Size is Medium from sheet 'Test 1' to a new Sheet called 'Mid market deals'
3. Copy the columns from test 1 that has a +33 country code to a new sheet called 'France Deals'
4. Plot a graph based on the 'France Deals' sheet on the sales value against Year id.

1. If Phone column in ‘Test 1’ doesn’t have a country code that starts with ‘+x' or ‘(xyz)’ where x, y,z are numbers, then add +1 as the country code.
2. Move the rows from ‘Test 1’ that has a ‘+1’ country code in phone column to a new sheet called ‘US Deals’ 
3. Make graph showing the cumulative sales value from the US deals sheet and plot it against the year id.

Pre:
1. Fill empty cells with n/a
2. 

1. Create a new column in ’Test 1’ by writing excel formulas that has a boolean output. If the deal size is ‘Medium’, Phone number starts with ‘+33’ and the YEAR_ID is either 2003 or 2004, then the formula output should be True, else the output should be false. The name of the column is ‘Relavancy’.
2. Make a graph on the entries from ‘Test 1’ whose Relevancy is True that shows the cumulative sales against year_id.

1. Create a new sheet called Formatted source.
2. Use the entries in the third row in Test 1 sheet as column names for the Formatted source sheet, except for the first entry in that row.
3. Add ‘Project count’ as the last column name in Formatted source.
4. Iterate through the entries in Test 1 sheet and store the last non-empty value in column 1 as ‘last project count’ variable. If the first column value in the row is empty, second column value is not ‘Unique id’ and the row is not empty with no data, then copy the row from the second cell to the last one and add it as an entry to Formatted source. Also add the value of ‘last project count’ as the entry to the project count column in the sheet.

1. Copy the third row in Test 1 and use them as the column names for test 1 by inserting a new first row. Also create a new column in the end called ‘Project’.
2.  Remove the rows that have ‘Unique ID’ in the second column excluding the first row.
3. Iterate through every row in Test 1 and store the last encountered non-null value in the first column as project name variable and add a new cell in the row’s end with the project name variable as the cell’s value.

1. Remove all empty rows.
2. Iterate through every row in test 1 and copy the last non-null value of the first column into that row’s project column’s value.

Step 1:
1. Copy the third row in test 1’s values and name the column’s in test 1 as those values
2. Delete all empty rows.
3. Delete all rows whose first cell is empty and second cell’s value is ‘Unique id’ excluding the first row.
4. Rename the first column in test 1 as Projects.

Step 2:
1. Rename the first column in test 1 as Projects.
2. Iterate through the cells in the projects column in the test 1 sheet and copy the cell value if it is not empty and paste it in the following empty cells, until another non-empty cell is detected then copy the new value to paste in the following empty cells.
3. Delete the rows in the test 1 sheet where the first cell is non-empty but all the following cells are empty.

1. Copy the third row in test 1’s values and name the column’s in test 1 as those values
2. Delete all empty rows.
3. Delete all rows whose first cell is empty and second cell’s value is ‘Unique id’ excluding the first row.
4. Rename the first column in test 1 as Projects.
5. Rename the first column in test 1 as Projects.
6. Iterate through the cells in the projects column in the test 1 sheet and copy the cell value if it is not empty and paste it in the following empty cells, until another non-empty cell is detected then copy the new value to paste in the following empty cells.
7. Delete the rows in the test 1 sheet where the first cell is non-empty but all the following cells are empty.
"""